import xlrd,xlwt
import csv
from openpyxl import Workbook,load_workbook
from xlutils.copy import copy
from kpi.models import Themetics,Userslist,Indicators,IndicatorTargetAchievements,Urban,IndicatorTargets
from farmers.models import Farmers ,LandDetails

""" Writing  a xlsv file  through importing a data from themetic model"""

def Themetics_xlsx(path):
    themetic=Themetics.objects.all()
    xls_file=xlrd.open_workbook(path)
    total_sheets = xls_file.nsheets
    wb=copy(xls_file)
    worksheet_data=wb.add_sheet("list_of_themetics")
    worksheet_data.write(0,0,"Themetic Id")
    worksheet_data.write(0,1,"Active")
    worksheet_data.write(0,2,"Themetic Name")
    worksheet_data.write(0,3,"Themetic Code")
    for i in range(len(themetic)):
        worksheet_data.write(i+1,0,themetic[i].id)
        worksheet_data.write(i+1,1,themetic[i].active)
        worksheet_data.write(i+1,2,themetic[i].themeticname )
        worksheet_data.write(i+1,3,themetic[i].code )
    wb.save("/home/mahiti/Desktop/xlsxfiles/themetics.xlsx")
path="/home/mahiti/Desktop/xlsxfiles/themetics.xlsx"






"""exporting a xlsx file's data into a database"""


def exporting_userdata(path2):
    wb=xlrd.open_workbook(path2)
    sheet=wb.sheet_by_index(1)
    data_length=sheet.nrows
    for i in range(1,data_length):
        user=sheet.row_values(i,0)
        Userslist.objects.create(user=user[0],username=user[1],firstname=user[2],lastname=user[3],roleid=user[4],roletype=user[5])
path2="/home/mahiti/Desktop/xlsxfiles/user_data.xlsx"



""" Exporting a data from model Indicators to a xlsx sheet """


def indicators_xlsx(path3):
    #create a workbook
    wb=Workbook()
    ws=wb.active
    ws.title="Indicators Data"
    wb.save(path3)

    #opening a saved workbook
    wb=load_workbook(path3)
    ws=wb.active
    column_name=("id","active","created","modified","indicatorname","shortcode","themetic id")
    ws.append(column_name)
    wb.save(path3)
    Indicator_list=Indicators.objects.all()

    for i in Indicator_list:
        row_data=(i.id , i.active, i.created , i.modified , i.indicatorname,i.shortcode ,i.themetic_id)
        ws.append(row_data)
        wb.save(path3)
path3="/home/mahiti/Desktop/xlsxfiles/indicator.xlsx"





""" Exporting a farmers details  from databse in two diffrent excel sheet"""

def farmer_xlsx(path4):
    wb=Workbook()
    wb.create_sheet("Farmers_details")
    wb.create_sheet("Land_details")
    wb.save(path4)
    
    open_wb=load_workbook(path4)

    farmer_sheet=wb["Farmers_details"]
    farmer_column=("Id","Name","Age","Address","ContactNo","Number of Family Member")
    farmer_sheet.append(farmer_column)

    land_sheet=wb["Land_details"]
    land_column=("Id","Farmer Name","Farmer id","Land location","Area in acres","Prodution in tonnes","Crop name")
    land_sheet.append(land_column)

    farmer_details=Farmers.objects.all()
    for farmer in farmer_details:
        farmer_column=(farmer.id,farmer.name,farmer.age,farmer.Address,farmer.contactno,farmer.familymember)
        farmer_sheet.append(farmer_column)

        land_detail=farmer.theme.all()

        for land_data in land_detail:
            land_column=(land_data.id,farmer.name,land_data.farmer_id,land_data.location,land_data.landarea,land_data.annualproduction,land_data.cropname)
            land_sheet.append(land_column)
    wb.save(path4)
   
path4 ="/home/mahiti/Desktop/xlsxfiles/farmers.xlsx"





"""Writing the csv file inot the table using pandas libr"""

def indicatorachievedtarget_csv(path5):
    with open(path5,"wb") as indicatorachievedtarget:
        achievedtarget=csv.writer(indicatorachievedtarget,delimiter=',')
        achievedtarget.writerow(["Active" ,"Created","Modified","Themeticname","Indicatortarget_name","Expected target","Achievedtarget"])
        targets=IndicatorTargetAchievements.objects.all()
        for target in targets:
            achievedtarget.writerow([target.active,target.created,target.modified,target.indicatortarget.indicator.themetic.themeticname,target.indicatortarget.indicator.indicatorname,target.indicatortarget.target,target.achievedtarget])
path5="/home/mahiti/Desktop/xlsxfiles/achievedtarget.csv"



def exporting_urbancsv_to_database(path6):
    with open(path6,"r") as urban_data:
        urban_data=csv.reader(urban_data)
        urban_data.next()
        for i in urban_data:
            Urban.objects.create(country=i[0],state=i[1],district=i[2],city=i[3],area=i[4],ward=i[5],mohala=i[6],geographicalarea=i[7])

path6="/home/mahiti/Desktop/xlsxfiles/cry_Urban_data.csv"


def models_to_xlsx(model,file_path,title):
    wb=Workbook()
    ws=wb.active
    ws.title=title
    fields_list=[ i.name for i in model._meta.fields]
    ws.append(fields_list)
    indicators_detail=model.objects.all()
    for detail in indicators_detail:
        indi_data=[]
        instance=model.objects.get(pk=detail.pk).__dict__
        for i in fields_list:
            if i+"_id" in instance:
                indi_data.append(instance[i+"_id"])
            else:
                indi_data.append(instance[i])
        ws.append(indi_data)
    wb.save(file_path)
  

file_path="/home/mahiti/Desktop/indicator_achieved_data_list.xlsx"


def insert_to_database(model,field_dic):
    model.objects.create(**field_dic)

def xlsx_to_models(model,file_path,sheet_name):
    from collections import OrderedDict
    open_wb=load_workbook(file_path)
    ws=open_wb[sheet_name]
    row= ws.max_row
    column=ws.max_column
    fields_dic=OrderedDict()
    fields_list=[str(ws.cell(row=1,column=i).value)for i in range(1,column+1)]
    for i in fields_list:
        fields_dic[i]=""
    for r in range(2,row+1):
        value=[]
        for c in range(1,column+1):
            cell=ws.cell(row=r,column=c)
            value.append(str(cell.value))
        
        index=0
        for i in fields_dic:
            fields_dic[i]=value[index]
            index+=1
        insert_to_database(model,fields_dic)
file_path="/home/mahiti/Desktop/xlsxfiles/user_data_list.xlsx"



    


















